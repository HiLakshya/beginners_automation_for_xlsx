import os
import openpyxl as xl

# Specify the directory containing your XLSX files
directory_path = os.getcwd()

# Iterate over each file in the directory
for filename in os.listdir(directory_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(directory_path, filename)

        wb = xl.load_workbook(file_path)
        sheet = wb['Sheet1']

        print(f"File: {filename}, Column A:")
        cell_a1 = sheet['A1']
        print(cell_a1.value)

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 1)
            print(cell.value)

        print("\n")

        print(f"File: {filename}, Column B:")
        cell_b1 = sheet['B1']
        print(cell_b1.value)

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            print(cell.value)

        print("\n")
        print("Updating Marks: Subtracting 5\n")

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            cell.value = cell.value - 5

        

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            print(cell.value)


        modified_file_path = os.path.join(directory_path, f"Modified_{filename}")
        wb.save(modified_file_path)
        print(f"File '{filename}' modified and saved as '{modified_file_path}'\n")
