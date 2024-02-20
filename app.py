import openpyxl as xl
from openpyxl.chart import BarChart, Reference 

wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1'] # Navigating to cell A1
print(cell.value) 

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    print(cell.value)

print("\n")

print(sheet['b1'].value)  # Navigating to cell B1
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    print(cell.value)

print("\n")
print("Updating Marks : Subtracting 5\n")

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    cell.value = cell.value - 5

wb.save('Book2.xlsx')  # Saving the file

print("After Updating Marks : Subtracting 5\n")

wb = xl.load_workbook('Book2.xlsx')
print(wb['Sheet1']['b1'].value)  # Navigating to cell B1
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    print(cell.value)

print("\n")

sheet = wb['Sheet1']
# Adding chart to the file
values = Reference(sheet, min_col=2, min_row=2, max_col=2, max_row=sheet.max_row)
chart = BarChart()
chart.add_data(values)

# Add the chart to the sheet
sheet.add_chart(chart, 'C2')

# Save the workbook after adding the chart
wb.save('Book2.xlsx')
